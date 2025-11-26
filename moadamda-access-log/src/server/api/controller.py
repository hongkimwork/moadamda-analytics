import pandas as pd
import openpyxl
from openpyxl.styles import Alignment
import imgkit
from PIL import Image
import io
import json
from datetime import datetime, timedelta
from conf.settings import settings
from server.api.model import Model
from slack_sdk import WebClient
from slack_sdk.errors import SlackApiError
from core.postgres import connection
import traceback
import asyncio
from core.cafe24 import Cafe24
from core.ga4 import GA4
from common import utils


class Controller():
    def __init__(self, pg):
        self.pg = pg
        self.model = Model(pg)
        self.cafe24 = Cafe24()
        self.ga4 = GA4()
    async def test(self):
        await self.process_slack_message(
            '1759386243.562649',
            """:gift::gift::gift: 주문접수알림 :gift::gift::gift:\n결제수단 : prepaid\n주문번호 : 20251002-0000102\n제      품 : ★리뉴얼★ [반값특가] 건강을 모아담다\n주문금액 : 55000.00원\n결제금액 : 55000.00원\n주문자/수령자 : 김현희, 김현희', 'team': 'T022G500A2K', 'bo
t_profile': {'id': 'B05GH0EAN80', 'deleted': False, 'name': 'Cafe24 Recipe Official', 'updated': 1688645086, 'app_id': 'A0188THGG5B', 'icons': {'image_36': 'https://avatars.slack-edge.com/2020-08-04/1275582639174_61f44f26cadff4f48a76_36.jpg', 'image_48': 'https://avatars.slac
k-edge.com/2020-08-04/1275582639174_61f44f26cadff4f48a76_48.jpg', 'image_72': 'https://avatars.slack-edge.com/2020-08-04/1275582639174_61f44f26cadff4f48a76_72.jpg'}, 'team_id': 'T022G500A2K'}, 'blocks': [{'type': 'rich_text', 'block_id': 'SQL', 'elements': [{'type': 'rich_tex
t_section', 'elements': [{'type': 'emoji', 'name': 'gift', 'unicode': '1f381'}, {'type': 'emoji', 'name': 'gift', 'unicode': '1f381'}, {'type': 'emoji', 'name': 'gift', 'unicode': '1f381'}, {'type': 'text', 'text': ' 주문접수알림 '}, {'type': 'emoji', 'name': 'gift', 'unicode
': '1f381'}, {'type': 'emoji', 'name': 'gift', 'unicode': '1f381'}, {'type': 'emoji', 'name': 'gift', 'unicode': '1f381'}, {'type': 'text', 'text': '\n결제수단 : prepaid\n주문번호 : 20251002-0000102\n제      품 : ★리뉴얼★ [반값특가] 건강을 모아담다\n주문금액 : 55000.00원\n결
제금액 : 55000.00원\n주문자/수령자 : 김현희, 김현희""",'C05AGK50VNY'
        )
        self.cafe24 = Cafe24()
        a = await self.cafe24.get_order_info('20240703-0000014')
        member_id = a['order']['member_id']
        b = await self.cafe24.get_buyer_info('20240703-0000014')
        c = await self.cafe24.get_member_info(member_id)
        end_date = datetime.now()
        mid_date = end_date - timedelta(days=90)
        start_date = mid_date - timedelta(days=90)
        d = await self.cafe24.get_member_order_history(
            member_id=member_id,
            start_date=start_date.strftime('%Y-%m-%d'),
            end_date=mid_date.strftime('%Y-%m-%d')
        )
        f = await self.cafe24.get_member_order_history(
            member_id=member_id,
            start_date=mid_date.strftime('%Y-%m-%d'),
            end_date=end_date.strftime('%Y-%m-%d')
        )

        print('a')
    async def add_access_log(
        self, url, referer,user_agent, client_ip, cookies, log_type, device, navigation_type
    ):
        if 'crema-product-reviews' in url:
            url = '&'.join([row for row in url.split('&') if 'crema-product-reviews' not in row])
        unquote_url = utils.ununquote(url)
        unquote_referer = utils.ununquote(referer)
        if log_type == 'enter':
            await self.model.add_access_log(
                unquote_url,unquote_referer,user_agent,client_ip,cookies, device, navigation_type
            )
        else:
            try:
                cookies_dict = json.loads(cookies)
                mduuid = cookies_dict['_mdUUID']
                log_list = await self.model.get_access_log(referer, unquote_referer, mduuid)
                for row in log_list:
                    if row and (row['url'] == url or row['url'] == unquote_url):
                        await self.model.update_access_log(row['idx'],cookies)
                        return
                else:
                    await self.model.add_access_log(
                        unquote_url,unquote_referer,user_agent,client_ip,cookies, device, navigation_type
                    )
                    print(datetime.now())
                    print(url)
                    print(mduuid)
                    return
            except Exception as e:
                traceback.print_exc()
    async def add_error_log(
        self, url, message,source,lineno,colno,stack
    ):
        await self.model.add_error_log(url, message, source, lineno, colno, stack)

    async def get_log(
        self, start_date, end_date
    ):

        response = {}
        fbclid_set = set()
        start_date = datetime.strptime(start_date,'%Y-%m-%d')
        end_date = datetime.strptime(end_date,'%Y-%m-%d')
        for n in range((end_date - start_date).days):
            current_date = start_date + timedelta(days=n)
            response[current_date.strftime('%Y-%m-%d')] = [{} for _ in range(24)]
        log_list = await self.model.get_access_log_for_analysis(start_date, end_date)
        key_set = set()
        for row in log_list:
            row = dict(row)
            fbclid = row.pop('fbclid','')
            if not fbclid:
                fbclid = row['ip']+row['url']
            if fbclid in fbclid_set:
                continue
            fbclid_set.add(fbclid)
            row['utm_content'] = (row.get('utm_content','') or '').replace('+',' ')
            key = f'{row["utm_source"]}{row["utm_campaign"]}{row["utm_medium"]}{row["utm_content"]}'
            access_date = row['access_date'].strftime('%Y-%m-%d')
            access_hour = row['access_date'].hour
            if key not in key_set:
                for d,hour_list in response.items():
                    for data in hour_list:
                        data[key] = {
                            'utm_source' : row['utm_source'],
                            'utm_campaign' : row['utm_campaign'],
                            'utm_medium' : row['utm_medium'],
                            'utm_content' : row['utm_content'],
                            'count' : 0
                        }
            key_set.add(key)
            response[access_date][access_hour][key]['count'] += 1
        for d,hour_list in response.items():
            response[d] = [list(row.values()) for row in hour_list]
            pass
        return response

    async def add_options(self):
        return
        UNIT_COUNT = 100
        offset = 0
        limit = UNIT_COUNT
        for _ in range(10):
            product_list = await self.cafe24.get_product_list(offset,limit)
            if len(product_list['products']) < UNIT_COUNT:
                break
            offset += UNIT_COUNT
            limit += UNIT_COUNT
        for product_info in product_list['products']:
            for options in product_info['options']['additional_options']:
                if options['additional_option_name'] == settings.SESSION_OPTION_NAME:

                    r = await self.cafe24.add_product_additional_options(
                        product_info['product_no'],
                        [
                            {
                                'additional_option_name' : settings.SESSION_OPTION_NAME,
                                'required_additional_option' : 'F',
                                'additional_option_text_length' : 50,
                            }
                        ]
                    )
                    break
            else:
                r = await self.cafe24.add_product_additional_options(
                    product_info['product_no'],
                    [
                        {
                            'additional_option_name' : settings.SESSION_OPTION_NAME,
                            'required_additional_option' : 'F',
                            'additional_option_text_length' : 50,
                        }
                    ]
                )
                pass
    async def send_status(
        self,
    ):

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(['날짜','아이피','이름','구매 금액','환경','채널','캠페인','이전 구매 기록','동일 쿠키 유입경로','동일 아이피 유입경로','이동경로'])
        async with connection() as pg:
            self.pg = pg
            self.model = Model(pg)
            self.cafe24 = Cafe24()
            start_date = datetime.now() - timedelta(days=1)
            end_date = datetime.now()
            order_list = await self.model.get_order_success_list(
                datetime(start_date.year, start_date.month, start_date.day),
                datetime(end_date.year, end_date.month, end_date.day),
            )
            for url in order_list:
                try:
                    order_id = url['url'].split('order_id=')[1]
                    (
                        order_info,
                        user_data,
                        access_date,
                        ip,
                        device,
                        name,
                        order_price,
                        order_history,
                        ip_campaign_history,
                        uuid_campaign_history,
                        page_move_history
                    ) = await self._get_order_info(order_id)
                    max_date = ''
                    max_data = ''
                    if len(ip_campaign_history) > 1:
                        max_date = ip_campaign_history[-1][0]
                        max_data = ip_campaign_history[-1][1]
                    if len(uuid_campaign_history) > 1:
                        _max_date = uuid_campaign_history[-1][0]
                        _max_data = uuid_campaign_history[-1][1]
                        if not max_date or datetime.strptime(max_date,'%Y-%m-%d %H:%M:%S') < datetime.strptime(_max_date,'%Y-%m-%d %H:%M:%S'):
                            max_date = _max_date
                            max_data = _max_data
                    utm_source = ''
                    utm_campaign = ''
                    if max_data:
                        utms = max_data.split('/')
                        utm_source = utms[0]
                        utm_campaign = utms[1]
                    sheet.append([
                        access_date,ip,
                        name,order_price,
                        device,utm_source,utm_campaign,
                        '\n'.join(['    '.join([str(col) for col in row]) for row in order_history[1:]]),
                        '\n'.join(['    '.join([str(col) for col in row]) for row in uuid_campaign_history[1:]]),
                        '\n'.join(['    '.join([str(col) for col in row]) for row in ip_campaign_history[1:]]),
                        '\n'.join(['    '.join([str(col) for col in row]) for row in page_move_history[1:]])
                    ])
                except:
                    pass

        # 줄바꿈 설정
        for col in ['H', 'I', 'J', 'K']:  # '동일 쿠키 유입경로', '동일 아이피 유입경로' 컬럼 인덱스
            col_idx = openpyxl.utils.cell.column_index_from_string(col)
            for row in range(2, len(order_list) + 1):
                cell = sheet.cell(row=row, column=col_idx)
                cell.alignment = Alignment(wrap_text=True)

        # 컬럼 너비 자동 조정 (선택사항)
        col_widthes = [19,15,10,10,10,10,20,100,100,100,100]
        i = 0
        for col in sheet.columns:
            column = col[0].column_letter  # Get the column name
            sheet.column_dimensions[column].width = col_widthes[i]
            i += 1

        # 모든 셀 상하, 좌우 가운데 정렬 설정
        alignment_hor_center = Alignment(horizontal='center', vertical='top', wrap_text=True)
        alignment_all_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        i = 0
        for row in sheet.iter_rows():
            for cell in row:
                if i == 0:
                    cell.alignment = alignment_hor_center
                else:
                    if cell.column_letter in ['H', 'I', 'J', 'K']:
                        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                    else:
                        cell.alignment = alignment_all_center
            i += 1
        file_path = f'{start_date.strftime('%Y-%m-%d')}_구매통계.xlsx'
        workbook.save(file_path)
        client = WebClient(token=settings.SLACK_BOT_TOKEN)
        image_res = client.files_upload_v2(
            channel=settings.SLACK_STATUS_CHANNEL_ID,
            file=file_path,
            title=file_path
        )
    async def process_slack_message(
        self,ts,text,channel
    ):
        async with connection() as pg:
            self.pg = pg
            self.model = Model(pg)
            self.cafe24 = Cafe24()
            if '주문접수알림' not in text and '무통장 입금 결제 완료' not in text:
                return
            order_id = ''
            for row in text.split('\n'):
                if '주문번호 : ' in row:
                    order_id = row.replace('주문번호 : ','')
                    break

            (
                order_info,
                user_data,
                access_date,
                ip,
                device,
                name,
                order_price,
                order_history,
                ip_campaign_history,
                uuid_campaign_history,
                page_move_history
            ) = await self._get_order_info(order_id)

            if (
                (
                    order_info['order']['member_id'] == '' and order_info['order']['market_id'] in ['NCHECKOUT','KAKAOPAY']
                )
                or '무통장 입금 결제 완료' in text
            ):
                event_name = ''
                if order_info['order']['market_id'] == 'NCHECKOUT':
                    event_name = 'purchase_npay_go'
                elif order_info['order']['market_id'] == 'KAKAOPAY':
                    event_name = 'purchase_kpay_go'
                else:
                    event_name = 'purchase_cash_go'
                if '_ga' in user_data['cookies']:
                    r = await self.ga4.send_purchase_info(
                        user_data['cookies']['_ga'],
                        event_name,
                        order_info['order']
                    )
                else:
                    print('_ga 값 없음')
            if '무통장 입금 결제 완료' in text:
                client = WebClient(token=settings.SLACK_BOT_TOKEN)
                chat_res = client.chat_postMessage(
                    channel=channel,
                    thread_ts=ts,
                    text='무통장입금 GA 전송 완료'
                )
                return
            clck = user_data['cookies'].get('_clck','')
            if clck:
                clarity_id = clck.split('|')[0]
            else:
                clarity_id = ''
            await self.make_image(
                access_date,
                ip, device,
                order_price,
                order_history,
                ip_campaign_history,
                uuid_campaign_history,
                page_move_history
            )
            blocks = await self.make_slack_block(
                access_date,
                ip, device,
                order_price,
                order_history,
                ip_campaign_history,
                uuid_campaign_history,
                page_move_history,
                clarity_id
            )
            try:
                client = WebClient(token=settings.SLACK_BOT_TOKEN)
                image_res = client.files_upload_v2(
                    channel=channel,
                    file='purchase_info.png',
                    title='구매정보 분석',
                    thread_ts=ts
                )
                chat_res = client.chat_postMessage(
                    channel=channel,
                    thread_ts=ts,
                    text='분석완료',
                    blocks=blocks
                )
            except SlackApiError as e:
                traceback.print_exc()

    async def _get_order_info(self, order_id):
        order_info = await self.cafe24.get_order_info(order_id)
        order_price = "{:,}".format(int(order_info['order']['initial_order_amount']['order_price_amount'].split('.')[0]))
        uuid = ''
        for item in order_info['order']['items']:
            for option_row in item['additional_option_values']:
                option_name = option_row['name']
                option_value = option_row['value']
                if option_name == 'additional_options' and '=' in option_value: #네이버페이
                    options = option_value.split('&')
                    for option in options:
                        option_name = option.split('=')[0]
                        option_value = option.split('=')[1]
                        if option_name == settings.SESSION_OPTION_NAME:
                            uuid = option_value
                else:
                    if option_name == settings.SESSION_OPTION_NAME:
                        uuid = option_value
        for i in range(3):
            if uuid:
                user_data = await self.model.get_user_info_from_uuid(uuid)
            else:
                user_data = await self.model.get_user_info_from_order_id(order_id)
            if not user_data and i == 2 and order_info['order']['member_id']:
                user_data = await self.model.get_user_info_from_member_id(order_info['order']['member_id'])
            if not user_data:
                await asyncio.sleep(i+1)
                continue
            else:
                user_data = dict(user_data)
                user_data['cookies'] = json.loads(user_data['cookies'])
                break
        ip = user_data['ip']
        uuid = user_data['uuid']
        name = order_info['order']['billing_name']
        access_date = user_data['access_date']
        access_date.strftime("%Y-%m-%d %H:%M:%S")
        device = user_data['device']

        order_history = await self._get_order_history(order_info['order']['member_id'],order_id)
        ip_history = await self.model.get_access_log_from_ip(ip)
        ip_campaign_history = await self._get_campaign_history(ip_history)
        uuid_history = await self.model.get_access_log_from_uuid(uuid)
        uuid_campaign_history = await self._get_campaign_history(uuid_history)
        page_move_history = await self._get_page_move_history(uuid_history)
        return (
            order_info,
            user_data,
            access_date,
            ip,
            device,
            name,
            order_price,
            order_history,
            ip_campaign_history,
            uuid_campaign_history,
            page_move_history
        )
    async def _get_order_history(self, member_id, order_id):
        res = [['날짜','상품명','금액']]
        if not member_id:
            return res
        try:
            order_id_set = set()
            order_id_set.add(order_id)
            raw_order_history = []
            start_date = datetime.now() - timedelta(days=360)
            for i in range(4):
                end_date = start_date + timedelta(days=90)
                for _ in range(3):
                    try:
                        orders = await self.cafe24.get_member_order_history(
                            member_id=member_id,
                            start_date=start_date.strftime('%Y-%m-%d'),
                            end_date=end_date.strftime('%Y-%m-%d')
                        )
                        raw_order_history.extend(reversed(orders['orders']))
                        break
                    except:
                        await asyncio.sleep(0.5)
                        continue
                start_date = end_date
            for row in raw_order_history:
                if row['order_id'] in order_id_set:
                    continue
                order_id_set.add(row['order_id'])
                res.append(
                    [
                        row['order_date'].replace('T',' ')[:16],
                        ', '.join([item['product_name'] for item in row['items']]),
                        "{:,}".format(int(row['payment_amount'].split('.')[0]))
                    ]
                )
        except Exception as e:
            traceback.print_exc()
        return res
    async def _get_page_move_history(self, history):
        response = [['날짜','페이지','체류시간(초)']]
        for row in history:
            url = utils.ununquote(row['url']).replace('https://m.moadamda.com','').replace('https://moadamda.com','').split('?')[0]
            access_date = row['access_date']
            end_date = row['end_date']
            maintain_time = ''
            if end_date:
                maintain_time = (end_date-access_date).total_seconds()
            response.append([
                    access_date.strftime("%Y-%m-%d %H:%M:%S"),
                    url,
                    maintain_time
                ]
            )
        return response
    async def _get_campaign_history(self, history):
        already_link_set = set()
        response = [['날짜','캠페인 명']]
        for row in history:
            if 'utm_content' not in row['url'] or (row['navigation_type'] and row['navigation_type'] != '0' and row['url'] in already_link_set):
                continue
            already_link_set.add(row['url'])
            params = utils.ununquote(row['url']).split('?')[1].split('&')
            content = campaign = source = ''
            for param in params:
                if 'utm_source=' in param:
                    source = param.replace('utm_source=','')
                elif 'utm_content=' in param:
                    content = param.replace('utm_content=','')
                elif 'utm_campaign=' in param:
                    campaign = param.replace('utm_campaign=','')
            response.append(
                [
                    row['access_date'].strftime("%Y-%m-%d %H:%M:%S"),
                    f'{source}/{campaign}/{content}'
                ]
            )
        return response

    async def make_image(
        self,
        access_date,
        ip,device,
        order_price,
        order_history,
        ip_history,
        uuid_history,
        page_move_history
    ):
        data = {
            "구매 정보": [
                ["구매 날짜", access_date],
                ["아이피", ip],
                ["환경", device],
                ["구매 금액", order_price]
            ],
            "과거 구매 기록": order_history,
            "동일 쿠키 유입 기록": uuid_history,
            "동일 아이피 유입 기록": ip_history,
            "페이지 이동 경로": page_move_history
        }
        html = "<html><head><style>body {font-family: 'Nanum Gothic', 'Noto Sans CJK KR', sans-serif;} table {border-collapse: collapse; width: 100%; word-wrap: break-word;} th, td {border: 1px solid black; padding: 8px; text-align: left;} th {background-color: #f2f2f2;}</style></head><body>"

        for title, rows in data.items():
            df = pd.DataFrame(rows)
            html += f"<h2>{title}</h2>"
            html += df.to_html(index=False, header=False)

        html += "</body></html>"

        img = imgkit.from_string(html, False)
        image = Image.open(io.BytesIO(img))

        image.save("purchase_info.png")

    async def make_slack_block(
        self,
        access_date,
        ip,
        device,
        order_price,
        order_history,
        ip_history,
        uuid_history,
        page_move_history,
        clarity_id
    ):
        blocks = [
            {
                "type": "section",
                "text": {
                    "type": "mrkdwn",
                    "text": "Clarity"
                },
                'accessory':{
                    'type' : 'button',
                    'text': {
                        'type': 'plain_text',
                        'text' : '기록 영상 보러가기'
                    },
                    'url':f'https://clarity.microsoft.com/projects/view/m8rhk75ntf/impressions?UserId=is%3B{clarity_id}&date=Last%2030%20days',
                    'action_id' : 'button_click'
                }
            },
            {
                "type": "section",
                "text": {
                    "type": "mrkdwn",
                    "text": f"*구매 날짜:* {access_date}"
                }
            },
            {
                "type": "section",
                "text": {
                    "type": "mrkdwn",
                    "text": f"*아이피:* {ip}"
                }
            },
            {
                "type": "section",
                "text": {
                    "type": "mrkdwn",
                    "text": f"*환경:* {device}"
                }
            },
            {
                "type": "section",
                "text": {
                    "type": "mrkdwn",
                    "text": f"*구매 금액:* {order_price}"
                }
            }
        ]
        blocks.append(
            {
                "type": "section",
                "text": {
                    "type": "mrkdwn",
                    "text": "*과거 구매 기록:*"
                }
            }
        )
        order_blocks = {
            "type": "context",
            "elements": [
                {
                    'type' : 'mrkdwn',
                    'text' : '\n'.join([f"`{row[0]}` `{row[1]}` `{row[2]}`" for row in order_history])
                }
            ]
        }
        blocks.append(order_blocks)
        blocks.append(
            {
                "type": "section",
                "text": {
                    "type": "mrkdwn",
                    "text": "*동일 쿠키 유입 기록:*"
                }
            }
        )
        uuid_blocks = {
            "type": "context",
            "elements": [
                {
                    'type' : 'mrkdwn',
                    'text' : '\n'.join([f"`{row[0]}` `{row[1]}`" for row in uuid_history])
                }
            ]
        }
        blocks.append(uuid_blocks)
        blocks.append(
            {
                "type": "section",
                "text": {
                    "type": "mrkdwn",
                    "text": "*동일 아이피 유입 기록:*"
                }
            }
        )
        ip_blocks = {
            "type": "context",
            "elements": [
                {
                    'type' : 'mrkdwn',
                    'text' : '\n'.join([f"`{row[0]}` `{row[1]}`" for row in ip_history])
                }
            ]
        }
        blocks.append(ip_blocks)
        blocks.append(
            {
                "type": "section",
                "text": {
                    "type": "mrkdwn",
                    "text": "*페이지 이동 경로:*"
                }
            }
        )
        page_move_blocks = {
            "type": "context",
            "elements": [
                {
                    'type' : 'mrkdwn',
                    'text' : '\n'.join([f"`{row[0]}` `{row[1]}` `{row[2]}`" for row in page_move_history])
                }
            ]
        }
        blocks.append(page_move_blocks)
        return blocks